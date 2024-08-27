from django.contrib import admin
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import tablib
from django.http import HttpResponse
from django.db.models import (
    Count,
    ExpressionWrapper,
    F,
    FloatField,
    Sum,
    Window,
)
import io
from .models import Contract
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
import zipfile
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.utils import timezone


@admin.action(description="Directors Approval Not Required")
def directors_approval_not_required(self, request, queryset):
    queryset.update(is_directors_approval="NO")


@admin.action(description="Directors Approval Required")
def directors_approval_required(self, request, queryset):
    queryset.update(is_directors_approval="YES")


@admin.action(description="Seamless Contract Updated")
def seamless_contract_updated(self, request, queryset):
    queryset.update(seamless_updated="YES")


@admin.action(description="Seamless Contract Not Updated")
def seamless_contract_not_updated(self, request, queryset):
    queryset.update(seamless_updated="NO")


@admin.action(description="Make Out of Contract")
def out_of_contract(self, request, queryset):
    queryset.update(is_ooc="YES")


@admin.action(description="Make Live")
def make_contract_live(self, request, queryset):
    queryset.update(contract_status="LIVE")


@admin.action(description="Pricing")
def make_contract_pricing(self, request, queryset):
    queryset.update(contract_status="PRICING")


@admin.action(description="Objection")
def make_contract_objection(self, request, queryset):
    queryset.update(contract_status="OBJECTION")


@admin.action(description="Locked")
def make_contract_locked(self, request, queryset):
    queryset.update(contract_status="LOCKED")


@admin.action(description="Contract Removed")
def contracts_removed(self, request, queryset):
    queryset.update(contract_status="REMOVED")


@admin.action(description="Contract Lost")
def contracts_lost(self, request, queryset):
    queryset.update(contract_status="LOST")


def bulk_quote_template(self, request, queryset):

    # Specify the fields you want to export
    fields_to_export = [
        "id",
        "client",
        "billing_address",
        "company_reg_number",
        "business_name",
        "site_address",
        "top_line",
        "mpan_mpr",
        "supplier",
        "contract_type",
        "contract_status",
        "contract_term",
        "contract_end_date",
        "eac",
        "is_directors_approval",
        "commission_per_annum",
        "commission_per_unit",
        "vat_rate",
    ]
    data = tablib.Dataset()
    data.headers = fields_to_export

    for obj in queryset:
        row = []
        for field in fields_to_export:
            if field == "contract_end_date":  # Check if the field is a date field
                date_value = getattr(obj, field)
                if date_value:  # Check if the date is not None
                    formatted_date = date_value.strftime("%d/%m/%Y")  # Format date to UK format
                    row.append(formatted_date)
                else:
                    row.append("")  # Append empty string if date is None
            elif field == "commission_per_unit":  # Check if the field is commission_per_unit
                commission_per_unit = getattr(obj, field)
                if commission_per_unit in [0.01, 0.02, 0.03]:
                    # Keep the value unchanged
                    converted_value = commission_per_unit
                else:
                    # Convert commission_per_unit to the desired format
                    converted_value = round(commission_per_unit * 100, 2)
                row.append(converted_value)
            else:
                row.append(getattr(obj, field))  # For non-date fields, just append the value
        data.append(row)

    response = HttpResponse(data.export("xlsx"), content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename="bulk_quote_template.xlsx"'
    return response


bulk_quote_template.short_description = "Bulk Quote Template"


def export_commissions_to_excel(self, request, queryset):
    # Check if the user has the 'can_export_commissions' permission
    # if not request.user.has_perm("contracts.can_export_commissions"):
    #     # If the user does not have the permission, display a message and redirect
    #     messages.error(request, "You do not have permission to export commissions.")
    #     return HttpResponseRedirect(reverse("admin:index"))

    # Create a BytesIO buffer
    buffer = io.BytesIO()

    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commissions by Client Contracts"

    # Code for adding data to the worksheet...
    columns = [
        "Client",
        "Contract Status",
        "Contract Type",
        "Originator",
        "Client Onboarded",
        "Supplier",
        "Utility",
        "MPAN/MPR",
        "Commission per Unit Rate",
        "Commission per Annum Rate",
        "Total EAC",
        "Number of Contracts",
        "Total Value per Contract",
    ]
    ws.append(columns)

    # Use the queryset directly which now reflects the user's selection
    selected_contracts = (
        queryset.values(
            "client__client",
            "contract_status",
            "contract_type",
            "client__originator",
            "client__client_onboarded",
            "supplier__supplier",
            "utility__utility",
            "mpan_mpr",
            "commission_per_unit",
            "commission_per_annum",
        )
        .annotate(
            total_eac=Sum("eac"),
            count=Count("id"),
            total_value_per_contract=ExpressionWrapper(
                F("total_eac") * F("commission_per_unit"),
                output_field=FloatField(),
            ),
        )
        .order_by("client__client", "utility__utility")
    )

    for contract in selected_contracts:
        ws.append(
            [
                contract["client__client"],
                contract["contract_status"],
                contract["contract_type"],
                contract["client__originator"],
                contract["client__client_onboarded"],
                contract["supplier__supplier"],
                contract["utility__utility"],
                contract["mpan_mpr"],
                contract["commission_per_unit"],
                contract["commission_per_annum"],
                contract["total_eac"],
                contract["count"],
                contract["total_value_per_contract"],
            ]
        )

    # Save the workbook to the BytesIO buffer
    wb.save(buffer)

    # Create the HttpResponse with the appropriate Excel mime type
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=contract_commissions_by_client.xlsx"

    return response


export_commissions_to_excel.short_description = "Export Client Commissions to Excel"


def export_expired_contracts(self, request, queryset):
    # Get today's date
    today = datetime.date.today()
    # Filter contracts that have expired prior to today
    contracts = (
        queryset.filter(
            contract_end_date__lt=today,
            is_ooc="YES",
        )
        .exclude(mpan_mpr__in=queryset.filter(contract_end_date__gte=today).values("mpan_mpr"))
        .order_by("client__client")
    )

    # Create a BytesIO buffer
    buffer = io.BytesIO()

    # Create Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expired Contracts"
    ws.append(["ID", "Client", "MPAN Number", "Contract Status", "Contract End Date", "OOC"])

    # Write contract data to Excel sheet with UK date format
    for contract in contracts:
        ws.append(
            [
                contract.id,
                contract.client.client,
                contract.mpan_mpr,
                contract.contract_status,
                contract.contract_end_date.strftime("%d/%m/%Y"),
                contract.is_ooc,
            ]
        )

    # Save workbook to BytesIO buffer
    wb.save(buffer)

    # Set response headers
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=expired_contracts_no_follow_on.xlsx"

    return response


export_expired_contracts.short_description = "Export expired contracts no follow on"


def export_corona_live_duplicates(self, request, queryset):
    # Fetch all duplicates with LIVE status and same contract end date
    duplicates_qs = (
        Contract.objects.filter(
            mpan_mpr__in=queryset.values_list("mpan_mpr", flat=True),
            contract_status="LIVE",
            supplier__supplier="Corona",
        )
        .exclude(mpan_mpr="11111")
        .annotate(
            row_number=Window(
                expression=Count("id"),
                partition_by=[
                    F("business_name"),
                    F("mpan_mpr"),
                    F("contract_end_date"),
                ],
                order_by=F("id").asc(),
            )
        )
        .filter(row_number__gt=1)
        .order_by("client__client")  # Sort by client name
    )

    # Prepare the Excel file
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Corona Live Duplicate Contracts"

    # Set headers with bold font
    headers = [
        "ID",
        "Client",
        "Business Name",
        "MPAN",
        "Contract Start Date",
        "Contract End Date",
        "Supplier",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    # Append data rows for each duplicate contract
    for contract in duplicates_qs:
        contract_start_date_uk = (
            contract.contract_start_date.strftime("%d/%m/%Y")
            if contract.contract_start_date
            else ""
        )
        contract_end_date_uk = (
            contract.contract_end_date.strftime("%d/%m/%Y") if contract.contract_end_date else ""
        )
        worksheet.append(
            [
                contract.id,
                contract.client.client,
                contract.business_name,
                contract.mpan_mpr,
                contract_start_date_uk,
                contract_end_date_uk,
                contract.supplier.supplier,
            ]
        )

    # Save and prepare for download
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="duplicate_contracts.xlsx"'
    return response


export_corona_live_duplicates.short_description = "Corona Live Duplicates"


def export_sse_live_duplicates(self, request, queryset):
    # Fetch all duplicates with LIVE status and same contract end date
    duplicates_qs = (
        Contract.objects.filter(
            mpan_mpr__in=queryset.values_list("mpan_mpr", flat=True),
            contract_status="LIVE",
            supplier__supplier="SSE",
        )
        .exclude(mpan_mpr="11111")
        .annotate(
            row_number=Window(
                expression=Count("id"),
                partition_by=[
                    F("business_name"),
                    F("mpan_mpr"),
                    F("contract_end_date"),
                ],
                order_by=F("id").asc(),
            )
        )
        .filter(row_number__gt=1)
        .order_by("client__client")  # Sort by client name
    )

    # Prepare the Excel file
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SSE Live Duplicate Contracts"

    # Set headers with bold font
    headers = [
        "ID",
        "Client",
        "Business Name",
        "MPAN",
        "Contract Start Date",
        "Contract End Date",
        "Supplier",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    # Append data rows for each duplicate contract
    for contract in duplicates_qs:
        contract_start_date_uk = (
            contract.contract_start_date.strftime("%d/%m/%Y")
            if contract.contract_start_date
            else ""
        )
        contract_end_date_uk = (
            contract.contract_end_date.strftime("%d/%m/%Y") if contract.contract_end_date else ""
        )
        worksheet.append(
            [
                contract.id,
                contract.client.client,
                contract.business_name,
                contract.mpan_mpr,
                contract_start_date_uk,
                contract_end_date_uk,
                contract.supplier.supplier,
            ]
        )

    # Save and prepare for download
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="duplicate_contracts.xlsx"'
    return response


export_sse_live_duplicates.short_description = "SSE Live Duplicates"

EXPORT_FIELDS = ["supplier", "client", "mpan_mpr", "business_name", "site_address"]


def get_field_value(obj, field):
    """Helper function to get string values from objects, including related objects"""
    value = getattr(obj, field)
    if field in ["supplier", "client"]:
        return str(value) if value else ""
    elif isinstance(value, bool):
        return "Yes" if value else "No"
    elif value is None:
        return ""
    return str(value)


def create_pdf(data, filename):
    """Create a PDF for a single row of data"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=50, bottomMargin=50, leftMargin=50, rightMargin=50
    )

    elements = []
    styles = getSampleStyleSheet()

    # Add title
    elements.append(Paragraph("Data Export Report", styles["Title"]))
    elements.append(Spacer(1, 20))

    # Add date
    current_date = timezone.now().strftime("%Y-%m-%d")
    elements.append(Paragraph(f"Date: {current_date}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Create table with field names and values
    table_data = [[field, data[field]] for field in EXPORT_FIELDS]
    table = Table(table_data, colWidths=[200, 300])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    return buffer


def export_to_excel_and_pdfs(modeladmin, request, queryset):
    # Excel export
    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Data"
    ws.append(EXPORT_FIELDS)
    for obj in queryset:
        ws.append([get_field_value(obj, field) for field in EXPORT_FIELDS])
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)

    # Create HTTP response
    response = HttpResponse(content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="exported_data.zip"'

    # Create a zip file
    with zipfile.ZipFile(response, "w") as zf:
        # Add Excel file to zip
        zf.writestr("exported_data.xlsx", excel_buffer.getvalue())

        # Create individual PDFs for each row
        for index, obj in enumerate(queryset):
            data = {field: get_field_value(obj, field) for field in EXPORT_FIELDS}
            pdf_buffer = create_pdf(data, f"data_{index + 1}.pdf")
            zf.writestr(f"data_{index + 1}.pdf", pdf_buffer.getvalue())

    return response


export_to_excel_and_pdfs.short_description = "Export to Excel and individual PDFs"

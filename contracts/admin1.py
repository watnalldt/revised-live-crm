import datetime
import tablib
from django.contrib import admin, messages
from django.db.models.functions import RowNumber

from .models import Contract
from clients.models import Client
from utilities.models import Supplier, Utility
from django.contrib.auth import get_user_model

# Required for links to client
from django.urls import reverse
from django.utils.html import format_html
from django.http import HttpResponse, HttpResponseRedirect

# Filters
from admin_auto_filters.filters import AutocompleteFilter
from rangefilter.filters import DateRangeFilter
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from django.db.models import F, Q, Count, Sum, FloatField, ExpressionWrapper, Window
from import_export import fields, resources
from import_export.admin import ImportExportModelAdmin
from import_export.widgets import ForeignKeyWidget
import openpyxl
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd

User = get_user_model()


# Custom Filters


class ClientFilter(AutocompleteFilter):
    title = "Client"  # display title
    field_name = "client"  # name of the foreign key field


class ClientManagerFilter(AutocompleteFilter):
    title = "Client Manager"  # display title
    field_name = "client_manager"  # name of the foreign key field


class SupplierFilter(AutocompleteFilter):
    title = "Supplier"  # display title
    field_name = "supplier"  # name of the foreign key field


class UtilityTypeFilter(AutocompleteFilter):
    title = "Utility Type"  # display title
    field_name = "utility"  # name of the foreign key field


class ContractFilter(admin.SimpleListFilter):
    title = "5% Vat no declaration"
    parameter_name = "contracts"

    def lookups(self, request, model_admin):
        return (("relevant_contracts", "Contracts 5% VAT no vat declaration"),)

    def queryset(self, request, queryset):
        if self.value() == "relevant_contracts":
            seven_days_ago = timezone.now() - timezone.timedelta(days=7)
            return queryset.filter(
                contract_start_date__lt=seven_days_ago, vat_rate="5%", vat_declaration_sent="NO"
            )


class SevenDaysAgoFilter(admin.SimpleListFilter):
    title = "Start Date More Than 7 Days Ago"
    parameter_name = "start_date_7_days_ago"

    def lookups(self, request, model_admin):
        return (
            ("yes", "Yes"),
            ("no", "No"),
        )

    def queryset(self, request, queryset):
        if self.value() == "yes":
            seven_days_ago = timezone.now() - timezone.timedelta(days=7)
            return queryset.filter(contract_start_date__lt=seven_days_ago)
        elif self.value() == "no":
            seven_days_ago = timezone.now() - timezone.timedelta(days=7)
            return queryset.exclude(contract_start_date__lt=seven_days_ago)
        return queryset


class StartDateIsNullFilter(admin.SimpleListFilter):
    title = _("Start Date Is Empty")
    parameter_name = "start_date_is_null"

    def lookups(self, request, model_admin):
        return (
            ("yes", _("Yes")),
            ("no", _("No")),
        )

    def queryset(self, request, queryset):
        if self.value() == "yes":
            return queryset.filter(contract_start_date__isnull=True)
        elif self.value() == "no":
            return queryset.filter(contract_start_date__isnull=False)


class MultiStatusFilter(admin.SimpleListFilter):
    title = _("Contract Status")  # Human-readable title which will be displayed
    parameter_name = "contract_status"  # URL parameter name for filtering

    def lookups(self, request, model_admin):
        """
        Returns a list of tuples. Each tuple contains a value and its corresponding
        human-readable name to be displayed in the admin filter sidebar.
        """
        # Assuming Contract.ContractStatus is a list or an enum-like of statuses; adapt if it's structured differently
        return [(status.value, status.name) for status in Contract.ContractStatus]

    def queryset(self, request, queryset):
        """
        Returns the filtered queryset based on the value(s) provided in the query string
        which are retrievable via `self.value()`. Supports multiple selections.
        """
        if self.value():
            # Splitting the values on commas allows for multiple statuses to be filtered
            filter_values = self.value().split(",")
            return queryset.filter(contract_status__in=filter_values)
        return queryset


class AccountManagerFilter(admin.SimpleListFilter):
    title = _('account manager')  # This is what will be displayed in the admin.
    parameter_name = 'account_manager'  # This is used in the URL for filtering.

    def lookups(self, request, model_admin):
        # Assuming 'AccountManager' is accessible via 'client__account_manager'.
        # Adjust the model path according to your model structure.
        AccountManager = model_admin.model.client.field.related_model.account_manager.field.model
        # Ensure to use `.distinct()` on the field that identifies account managers uniquely, typically their ID or email.
        # The following line assumes 'email' is a field on AccountManager that can serve as a unique identifier.
        return [(a.id, a.account_manager) for a in AccountManager.objects.order_by('account_manager').distinct('account_manager')]

    def queryset(self, request, queryset):
        # Filters the queryset based on the account manager's ID
        if self.value():
            return queryset.filter(client__account_manager__id=self.value())
        else:
            return queryset

# End Custom Filters


class ContractResource(resources.ModelResource):
    client = fields.Field(
        column_name="client", attribute="client", widget=ForeignKeyWidget(Client, "client")
    )
    client_manager = fields.Field(
        column_name="client_manager",
        attribute="client_manager",
        widget=ForeignKeyWidget(User, "email"),
    )
    supplier = fields.Field(
        column_name="supplier", attribute="supplier", widget=ForeignKeyWidget(Supplier, "supplier")
    )
    utility = fields.Field(
        column_name="utility", attribute="utility", widget=ForeignKeyWidget(Utility, "utility")
    )

    class Meta:
        model = Contract

        report_skipped = True
        import_id_fields = ("id",)
        export_order = [
            "id",
            "client",
            "client_group",
            "business_name",
            "site_address",
            "client_manager",
            "supplier",
            "utility",
            "mpan_mpr",
            "meter_serial_number",
            "meter_status",
            "smart_meter",
            "top_line",
            "account_number",
            "company_reg_number",
            "building_name",
            "billing_address",
            "is_directors_approval",
            "contract_type",
            "contract_status",
            "meter_onboarded",
            "vat_rate",
            "vat_declaration_sent",
            "vat_declaration_date",
            "vat_declaration_expires",
            "lock_in_date",
            "contract_start_date",
            "contract_end_date",
            "supplier_start_date",
            "contract_term",
            "is_ooc",
            "pence_per_kilowatt",
            "day_kilowatt_hour_rate",
            "night_rate",
            "annualised_budget",
            "day_consumption",
            "night_consumption",
            "contract_value",
            "standing_charge",
            "sc_frequency",
            "unit_rate_1",
            "unit_rate_2",
            "unit_rate_3",
            "supplier_coding",
            "seamless_status",
            "seamless_updated",
            "eac",
            "profile",
            "service_type",
            "feed_in_tariff",
            "kva",
            "commission_per_annum",
            "commission_per_unit",
            "commission_per_contract",
            "partner_commission",
            "notes",
        ]


class ContractAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    show_full_result_count = False
    resource_class = ContractResource
    list_per_page = 10
    ordering = ("id",)
    readonly_fields = ("commission_per_annum", "commission_per_unit")
    list_display = (
        "id",
        "contract_status",
        "business_name",
        "contract_type",
        "seamless_updated",
        "link_to_clients",
        "client_manager",
        "site_address",
        "supplier",
        "utility",
        "meter_serial_number",
        "mpan_mpr",
        "eac",
        "commission_per_annum",
        "commission_per_unit",
        "contract_start_date",
        "contract_end_date",
        "is_ooc",
        "is_directors_approval",
    )
    list_display_links = ("business_name",)
    list_filter = [
        "contract_type",
        "seamless_updated",
        MultiStatusFilter,
        ClientFilter,
        AccountManagerFilter,
        "client_group",
        ClientManagerFilter,
        SupplierFilter,
        UtilityTypeFilter,
        "seamless_status",
        "is_ooc",
        "is_directors_approval",
        ("contract_end_date", DateRangeFilter),
        ("contract_start_date", DateRangeFilter),
        "vat_rate",
        "vat_declaration_sent",
        StartDateIsNullFilter,
        "meter_status",
    ]
    fieldsets = (
        (
            "Site Information",
            {
                "description": "Enter the site details",
                "fields": (
                    (
                        "client",
                        "client_group",
                    ),
                    "business_name",
                    "site_address",
                    "client_manager",
                    "supplier",
                    "utility",
                    "mpan_mpr",
                    "meter_serial_number",
                    "meter_status",
                    "smart_meter",
                    "top_line",
                ),
            },
        ),
        (
            "Contract Information",
            {
                "description": "Contract Information",
                "fields": (
                    (
                        "account_number",
                        "company_reg_number",
                    ),
                    ("building_name", "billing_address"),
                    "is_directors_approval",
                    "contract_type",
                    "contract_status",
                    "meter_onboarded",
                ),
            },
        ),
        (
            "VAT Information",
            {
                "fields": (
                    "vat_rate",
                    "vat_declaration_sent",
                    "vat_declaration_date",
                    "vat_declaration_expires",
                )
            },
        ),
        (
            "Contract Date and Contract Rates",
            {
                "description": "Enter the following details",
                "fields": (
                    (
                        "lock_in_date",
                        "contract_start_date",
                        "contract_end_date",
                        "contract_term",
                    ),
                    (
                        "supplier_start_date",
                        "is_ooc",
                    ),
                    (
                        "pence_per_kilowatt",
                        "day_kilowatt_hour_rate",
                        "night_rate",
                        "annualised_budget",
                    ),
                    ("day_consumption", "night_consumption", "contract_value"),
                    ("standing_charge", "sc_frequency"),
                    ("unit_rate_1", "unit_rate_2", "unit_rate_3"),
                ),
            },
        ),
        (
            "Seamless Contract Information",
            {
                "description": "The following only applies to seamless contracts",
                "fields": (
                    ("supplier_coding",),
                    "seamless_status",
                    "seamless_updated",
                ),
            },
        ),
        (
            "Service Information",
            {
                "description": "Enter the following data",
                "fields": (
                    "eac",
                    "profile",
                    "service_type",
                    "feed_in_tariff",
                    "kva",
                ),
            },
        ),
        (
            "Commissions",
            {
                "description": "Commissions are applied at client level",
                "fields": (
                    "commission_per_annum",
                    "commission_per_unit",
                    "commission_per_contract",
                    "partner_commission",
                ),
            },
        ),
        ("Notes", {"description": "Additional Information", "fields": ("notes",)}),
    )
    autocomplete_fields = [
        "client",
        "client_manager",
        "supplier",
    ]

    # Custom Search for multiple mpans
    from django.db.models import Q

    def get_search_results(self, request, queryset, search_term):
        """
        Enhances filtering by allowing searches for multiple MPAN numbers separated by commas,
        and maintains compatibility with other applied filters.
        """
        use_distinct = False

        # Check if the search_term involves potential multiple MPAN numbers
        if "," in search_term:
            mpan_terms = [term.strip() for term in search_term.split(",") if term.strip()]
            if mpan_terms:
                q_objects = Q(mpan_mpr__iexact=mpan_terms[0])  # Start with the first term
                for term in mpan_terms[1:]:
                    q_objects |= Q(mpan_mpr__iexact=term)  # Use |= to add each subsequent term

                queryset = queryset.filter(q_objects).distinct()
                use_distinct = True
        else:
            # Handle single mpan_mpr or other search fields via superclass method
            queryset, use_distinct = super().get_search_results(request, queryset, search_term)

        return queryset, use_distinct

    search_fields = (
        "business_name",
        "client__client",
        "client_group",
        "utility__utility",
        "supplier__supplier",
        "mpan_mpr",
        "meter_serial_number",
        "site_address",
    )

    date_hierarchy = "contract_end_date"
    actions = [
        "bulk_quote_template",
        "export_combined_report",
        "export_commissions_to_excel",
        "export_duplicates_to_excel",
        # "export_expired_contracts",
    ]

    def bulk_quote_template(self, request, queryset):
        # Check if the user has the required permission
        if not request.user.has_perm("contracts.can_access_bulk_quote_template"):
            messages.error(request, "You do not have permission to use the bulk upload template.")
            # If the user does not have permission, return an HTTP Forbidden response or any other appropriate action
            return HttpResponseRedirect(reverse("admin:index"))

        # Specify the fields you want to export
        fields_to_export = [
            "id",
            "client",
            "billing_address",
            "company_reg_number",
            "business_name",
            "site_address",
            "top_line",
            "mpan_mpr",
            "supplier",
            "contract_status",
            "contract_end_date",
            "contract_term",
            "eac",
            "is_directors_approval",
            "commission_per_annum",
            "commission_per_unit",
        ]
        data = tablib.Dataset()
        data.headers = fields_to_export

        for obj in queryset:
            row = []
            for field in fields_to_export:
                if field == "contract_end_date":  # Check if the field is a date field
                    date_value = getattr(obj, field)
                    if date_value:  # Check if the date is not None
                        formatted_date = date_value.strftime("%d/%m/%Y")  # Format date to UK format
                        row.append(formatted_date)
                    else:
                        row.append("")  # Append empty string if date is None
                elif field == "commission_per_unit":  # Check if the field is commission_per_unit
                    commission_per_unit = getattr(obj, field)
                    if commission_per_unit in [0.01, 0.02, 0.03]:
                        # Keep the value unchanged
                        converted_value = commission_per_unit
                    else:
                        # Convert commission_per_unit to the desired format
                        converted_value = round(commission_per_unit * 100, 2)
                    row.append(converted_value)
                else:
                    row.append(getattr(obj, field))  # For non-date fields, just append the value
            data.append(row)

        response = HttpResponse(data.export("xlsx"), content_type="application/vnd.ms-excel")
        response["Content-Disposition"] = 'attachment; filename="bulk_quote_template.xlsx"'
        return response

    bulk_quote_template.short_description = "Bulk Quote Template"

    def export_combined_report(self, request, queryset):
        # Filter out contracts with statuses 'Lost' or 'Removed'
        queryset = queryset.exclude(contract_status__in=["LOST", "REMOVED"])

        # Set up the response for Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="All_Reports.xlsx"'

        # Use ExcelWriter to write multiple sheets
        with pd.ExcelWriter(response, engine="openpyxl") as writer:
            # Process and export "Contracts By Status"
            df_status = pd.DataFrame.from_records(
                queryset.annotate(
                    count=Count("id"), client_contract_status=F("contract_status")
                ).values("client_contract_status", "count")
            )
            pivot_table_status = df_status.pivot_table(
                values="count",
                index=["client_contract_status"],
                aggfunc="sum",
                margins=True,
                margins_name="Total",
            )
            pivot_table_status.to_excel(writer, sheet_name="Contract Status Count")

            # Apply a filter on the Excel sheet (for Excel 2010 and later)
            worksheet = writer.sheets["Contract Status Count"]
            worksheet.auto_filter.ref = "A:A"  # Apply auto filter on all columns

            # Process and export "DA Approval Status"
            df_approval = pd.DataFrame.from_records(queryset.values("is_directors_approval"))
            total_yes = (df_approval["is_directors_approval"] == "YES").sum()
            total_no = (df_approval["is_directors_approval"] == "NO").sum()
            approval_df = pd.DataFrame(
                {
                    "Approval Status": ["YES", "NO", "Total"],
                    "Total": [total_yes, total_no, total_yes + total_no],
                }
            )
            approval_df.to_excel(writer, index=False, sheet_name="DA Approval Status")

            # Process and export "OOC Status Counts"
            df_ooc = pd.DataFrame.from_records(queryset.values("is_ooc"))
            total_yes = (df_ooc["is_ooc"] == "YES").sum()
            total_no = (df_ooc["is_ooc"] == "NO").sum()
            ooc_df = pd.DataFrame(
                {
                    "OOC Status": ["YES", "NO", "Total"],
                    "Total": [total_yes, total_no, total_yes + total_no],
                }
            )
            ooc_df.to_excel(writer, index=False, sheet_name="OOC Status Count")

        return response

    export_combined_report.short_description = "Export Combined Report"

    def export_commissions_to_excel(self, request, queryset):
        # Check if the user has the 'can_export_commissions' permission
        if not request.user.has_perm("contracts.can_export_commissions"):
            # If the user does not have the permission, display a message and redirect
            messages.error(request, "You do not have permission to export commissions.")
            return HttpResponseRedirect(reverse("admin:index"))
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=commissions_by_client.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commissions by Client Contracts"

        columns = [
            "Client",
            "Contract Status",
            "Contract Type",
            "Originator",
            "Client Onboarded",
            "Utility",
            "Commission per Unit Rate",
            "Commission per Annum Rate",
            "Total EAC",
            "Number of Contracts",
            "Total Value per Contract",
        ]
        ws.append(columns)

        # Use the queryset directly which now reflects the user's selection
        selected_contracts = (
            queryset.values(
                "client__client",
                "contract_status",  # Assuming 'client' field is a ForeignKey to a Client model
                "contract_type",
                "client__originator",
                "client__client_onboarded",
                "utility__utility",  # Assuming 'utility' field is a ForeignKey to a Utility model
                "commission_per_unit",
                "commission_per_annum",
            )
            .annotate(
                total_eac=Sum("eac"),
                count=Count("id"),
                total_value_per_contract=ExpressionWrapper(
                    F("total_eac") * F("commission_per_unit"),
                    output_field=FloatField(),
                ),
            )
            .order_by("client__client", "utility__utility")
        )

        for contract in selected_contracts:
            ws.append(
                [
                    contract["client__client"],
                    contract["contract_status"],
                    contract["contract_type"],
                    contract["client__originator"],
                    contract["client__client_onboarded"],
                    contract["utility__utility"],
                    contract["commission_per_unit"],
                    contract["commission_per_annum"],
                    contract["total_eac"],
                    contract["count"],
                    contract["total_value_per_contract"],
                ]
            )

        wb.save(response)
        return response

    export_commissions_to_excel.short_description = "Export Commissions by to Excel"

    def export_duplicates_to_excel(self, request, queryset):
        # Filter contracts based on the provided queryset and annotate them with a count of duplicates
        contracts_with_duplicates = (
            Contract.objects.filter(mpan_mpr__in=queryset.values_list("mpan_mpr", flat=True))
            .values("mpan_mpr", "contract_end_date")
            .annotate(duplicates_count=Count("id"))
            .filter(duplicates_count__gt=1)
        )

        # Prepare a dictionary of MPAN/MPR and their corresponding contract_end_date with duplicates count
        duplicates = {
            item["mpan_mpr"]: {
                "contract_end_date": item["contract_end_date"],
                "duplicates_count": item["duplicates_count"],
            }
            for item in contracts_with_duplicates
        }

        # Filter original contracts including those in the duplicates dictionary
        annotated_contracts = Contract.objects.filter(mpan_mpr__in=duplicates.keys()).annotate(
            row_number=Window(
                expression=RowNumber(),
                partition_by=[F("mpan_mpr"), F("contract_end_date")],
                order_by=F("id").asc(),
            )
        )

        # Prepare the Excel file
        output, workbook = BytesIO(), Workbook()
        worksheet = workbook.active
        worksheet.title = "Duplicate Contracts"

        # Set headers with bold font, including duplicates count
        headers = [
            "ID",
            "Client",
            "Business Name",
            "MPAN",
            "Contract Status",
            "Contract Start Date",
            "Contract End Date",
            "Duplicates Count",
        ]
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # Append data rows for each duplicate contract, including duplicates count
        for contract in annotated_contracts:
            contract_start_date_uk = (
                contract.contract_start_date.strftime("%d/%m/%Y")
                if contract.contract_start_date
                else ""
            )
            contract_end_date_uk = (
                contract.contract_end_date.strftime("%d/%m/%Y")
                if contract.contract_end_date
                else ""
            )
            mpan_mpr = contract.mpan_mpr
            duplicates_count = (
                duplicates[mpan_mpr]["duplicates_count"] if mpan_mpr in duplicates else 0
            )
            worksheet.append(
                [
                    contract.id,
                    contract.client.client,
                    contract.business_name,
                    mpan_mpr,
                    contract.contract_status,
                    contract_start_date_uk,
                    contract_end_date_uk,
                    duplicates_count,  # Include duplicates count
                ]
            )

        # Save and prepare for download
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="duplicate_contracts.xlsx"'
        return response

    export_duplicates_to_excel.short_description = "Export Duplicate Contracts to Excel"

    def export_expired_contracts(self, request, queryset):
        # Filter contracts based on conditions
        contracts = queryset.filter(
            contract_end_date__lt=datetime.date(2024, 3, 31),
            is_ooc="YES",
        ).exclude(
            mpan_mpr__in=queryset.filter(contract_end_date__gt=datetime.date(2024, 3, 31)).values(
                "mpan_mpr"
            )
        )

        # Create Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Client", "MPAN Number", "Contract Status", "Contract End Date", "OOC"])

        # Write contract data to Excel sheet with UK date format
        for contract in contracts:
            ws.append(
                [
                    contract.id,
                    contract.client.client,
                    contract.mpan_mpr,
                    contract.contract_status,
                    contract.contract_end_date.strftime("%d/%m/%Y"),
                    contract.is_ooc,
                ]
            )

        # Set response headers
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=expired_contracts_no_follow_on.xlsx"

        # Save workbook to response
        wb.save(response)
        return response

    export_expired_contracts.short_description = "Export expired contracts no follow on"

    def link_to_clients(self, obj):
        link = reverse("admin:clients_client_change", args=[obj.client.id])
        return format_html(
            '<a href="{}">{}</a>',
            link,
            obj.client,
        )

    link_to_clients.short_description = "Clients"


admin.site.register(Contract, ContractAdmin)
